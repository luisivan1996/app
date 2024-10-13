from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify, Response
import psycopg2
import csv
import json
from fpdf import FPDF
import pandas as pd
import io
from werkzeug.security import generate_password_hash, check_password_hash
import xml.etree.ElementTree as ET
from datetime import date
from openpyxl import Workbook

from db import get_db_connection


app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Necesario para manejar sesiones

# Ruta para login
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM usuarios WHERE username = %s", (username,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and user[2] == password:  # Sin bcrypt por simplicidad
            session['user'] = user[1]
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos')

    return render_template('login.html')

# Ruta para dashboard
@app.route('/dashboard')
def dashboard():
    if 'user' in session:
        return render_template('dashboard.html')
    else:
        return redirect(url_for('login'))

# Ruta para listar alumnos
@app.route('/alumnos')
def listar_alumnos():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT * FROM estudiantes')
    alumnos = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('alumnos.html', alumnos=alumnos)

# Ruta para agregar alumno
@app.route('/alumnos/agregar', methods=['GET', 'POST'])
def agregar_alumno():
    if request.method == 'POST':
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        cedula = request.form['cedula']
        fecha_nacimiento = request.form['fecha_nacimiento']
        
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("INSERT INTO estudiantes (nombre, apellido, cedula, fecha_nacimiento) VALUES (%s, %s, %s, %s)",
                    (nombre, apellido, cedula, fecha_nacimiento))
        conn.commit()
        cur.close()
        conn.close()
        return redirect(url_for('listar_alumnos'))

    return render_template('agregar_alumno.html')

# Ruta para editar alumno
@app.route('/alumnos/editar/<int:id>', methods=['GET', 'POST'])
def editar_alumno(id):
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        cedula = request.form['cedula']
        fecha_nacimiento = request.form['fecha_nacimiento']
        cur.execute("UPDATE estudiantes SET nombre = %s, apellido = %s, cedula = %s, fecha_nacimiento = %s WHERE id = %s",
                    (nombre, apellido, cedula, fecha_nacimiento, id))
        conn.commit()
        return redirect(url_for('listar_alumnos'))
    
    cur.execute('SELECT * FROM estudiantes WHERE id = %s', (id,))
    alumno = cur.fetchone()
    cur.close()
    conn.close()

    return render_template('editar_alumno.html', alumno=alumno)

# Ruta para eliminar alumno
@app.route('/alumnos/eliminar/<int:id>')
def eliminar_alumno(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM estudiantes WHERE id = %s', (id,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for('listar_alumnos'))

# Rutas de Exportación
@app.route('/exportar/<formato>')
def exportar_datos(formato):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM estudiantes")
    estudiantes = cur.fetchall()
    column_names = [desc[0] for desc in cur.description]

    if formato == 'pdf':
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        for estudiante in estudiantes:
            pdf.cell(200, 10, txt=f"{estudiante}", ln=True)
        pdf_file = 'estudiantes.pdf'
        pdf.output(pdf_file)
        return send_file(pdf_file, as_attachment=True)

    elif formato == 'xlsx':
        df = pd.DataFrame(estudiantes, columns=column_names)
        excel_file = 'estudiantes.xlsx'
        df.to_excel(excel_file, index=False)
        return send_file(excel_file, as_attachment=True)

    elif formato == 'csv':
        df = pd.DataFrame(estudiantes, columns=column_names)
        csv_file = 'estudiantes.csv'
        df.to_csv(csv_file, index=False)
        return send_file(csv_file, as_attachment=True)

    elif formato == 'xml':
        root = ET.Element("Estudiantes")
        for estudiante in estudiantes:
            estudiante_element = ET.SubElement(root, "Estudiante")
            for i, col in enumerate(column_names):
                ET.SubElement(estudiante_element, col).text = str(estudiante[i])
        tree = ET.ElementTree(root)
        xml_file = 'estudiantes.xml'
        tree.write(xml_file)
        return send_file(xml_file, as_attachment=True)

    elif formato == 'json':
        data = []
        for est in estudiantes:
            estudiante_dict = {
                "ID": est[0],
                "Nombre": est[1],
                "Apellido": est[2],
                "Cedula": est[3],
                "Fecha de Nacimiento": est[4].strftime('%Y-%m-%d')  # Convertir la fecha a cadena
            }
            data.append(estudiante_dict)

        json_output = json.dumps(data, indent=4)
        return send_file(io.BytesIO(json_output.encode()), as_attachment=True, download_name='estudiantes.json', mimetype='application/json')

    return redirect(url_for('alumnos'))

# Ruta para generar reportes
@app.route('/reportes')
def reportes():
    return render_template('reportes.html')


# Cerrar sesión
@app.route('/logout')
def logout():
    session.pop('loggedin', None)
    session.pop('id', None)
    session.pop('username', None)
    flash('Has cerrado sesión correctamente.', 'success')
    return redirect(url_for('login'))

@app.route('/buscar_estudiantes')
def buscar_estudiantes():
    term = request.args.get('term')
    # Aquí debes tener tu conexión a la base de datos
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM estudiantes WHERE nombre ILIKE %s OR cedula ILIKE %s", (f'%{term}%', f'%{term}%'))
    estudiantes = cursor.fetchall()

    # Convertir a un formato JSON adecuado
    result = []
    for estudiante in estudiantes:
        result.append({
            'id': estudiante[0],
            'nombre': estudiante[1],
            'apellido': estudiante[2],
            'cedula': estudiante[3],
            'fecha_nacimiento': estudiante[4].isoformat()  # Convertir a formato ISO para JSON
        })

    return jsonify(result)

@app.route('/export/csv')
def exportar_csv():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM estudiantes WHERE nombre ILIKE %s OR cedula ILIKE %s", (f'%{term}%', f'%{term}%'))
    estudiantes = cursor.fetchall()

    # Crear CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Nombre', 'Apellido', 'Cédula', 'Fecha de Nacimiento'])
    for estudiante in estudiantes:
        writer.writerow(estudiante)

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), as_attachment=True, download_name='estudiantes.csv', mimetype='text/csv')


@app.route('/export/xlsx')
def exportar_xlsx():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM estudiantes WHERE nombre ILIKE %s OR cedula ILIKE %s", (f'%{term}%', f'%{term}%'))
    estudiantes = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Estudiantes"

    # Escribir encabezados
    sheet.append(['ID', 'Nombre', 'Apellido', 'Cédula', 'Fecha de Nacimiento'])

    for estudiante in estudiantes:
        sheet.append(estudiante)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='estudiantes.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/pdf')
def exportar_pdf():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM estudiantes WHERE nombre ILIKE %s OR cedula ILIKE %s", (f'%{term}%', f'%{term}%'))
    estudiantes = cursor.fetchall()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Reporte de Estudiantes", ln=True, align='C')

    for estudiante in estudiantes:
           pdf.cell(200, 10, txt=f"{estudiante[1]} {estudiante[2]} - Cédula: {estudiante[3]} - Fecha Nacimiento: {estudiante[4]}", ln=True)
    pdf_file = 'estudiantes.pdf'
    pdf.output(pdf_file)
    return send_file(pdf_file, as_attachment=True)


@app.route('/export/json')
def export_json():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM estudiantes WHERE nombre ILIKE %s OR cedula ILIKE %s", (f'%{term}%', f'%{term}%'))
    estudiantes = cursor.fetchall()

    result = [{'id': e[0], 'nombre': e[1], 'apellido': e[2], 'cedula': e[3], 'fecha_nacimiento': e[4].isoformat()} for e in estudiantes]

    response = json.dumps(result)
    return Response(response, mimetype='application/json', headers={"Content-Disposition": "attachment;filename=estudiantes.json"})


@app.route('/export/xml')
def exportar_xml():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM estudiantes WHERE nombre ILIKE %s OR cedula ILIKE %s", (f'%{term}%', f'%{term}%'))
    estudiantes = cursor.fetchall()

    root = ET.Element("estudiantes")
    for estudiante in estudiantes:
        est_elem = ET.SubElement(root, "estudiante")
        ET.SubElement(est_elem, "id").text = str(estudiante[0])
        ET.SubElement(est_elem, "nombre").text = estudiante[1]
        ET.SubElement(est_elem, "apellido").text = estudiante[2]
        ET.SubElement(est_elem, "cedula").text = estudiante[3]
        ET.SubElement(est_elem, "fecha_nacimiento").text = estudiante[4].isoformat()

    output = io.BytesIO()
    tree = ET.ElementTree(root)
    tree.write(output, encoding='utf-8', xml_declaration=True)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='estudiantes.xml', mimetype='application/xml')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
